from main import disnake, commands

from db_config import connect

import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


class Excel(commands.Cog):
    def __init__(self, bot):
        self.bot = bot

    # для выгрузки excel
    @commands.slash_command(name="get-users-excel", default_member_permissions=disnake.Permissions(administrator=True))
    @commands.is_owner()
    async def get_users_excel(self, interaction: disnake.ApplicationCommandInteraction):
        connection = connect()
        try:
            with connection.cursor() as cursor:
                cursor.execute("SELECT * FROM users")
                rows = cursor.fetchall()
                columns = [col[0] for col in cursor.description]  # Получаем названия столбцов
        finally:
            connection.close()

        df = pd.DataFrame(list(rows), columns=columns)

        wb = Workbook()
        ws = wb.active
        ws.title = "Users"

        header_font = Font(bold=True, color="000000")
        header_fill = PatternFill(start_color="ed5c5c", end_color="ed5c5c", fill_type="solid")

        # записываем заголовки
        for col_num, column_title in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num, value=column_title)
            cell.font = header_font
            cell.fill = header_fill

        # вносим строки
        for row_num, r in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
            ws.append(r)

            # проверка значения в столбце is_leader и закрашивание ячейки
            is_leader_index = df.columns.get_loc('is_leader')  # получаем индекс столбца is_leader
            if r[is_leader_index] == 1:
                ws.cell(row=row_num, column=is_leader_index).fill = PatternFill(start_color="5ced7c",
                                                                                end_color="5ced7c", fill_type="solid")

        # Сохранение файла
        wb.save("data/users.xlsx")

        await interaction.response.send_message("Файл *users.xlsx* выгружен", ephemeral=True)


def setup(bot):
    bot.add_cog(Excel(bot))
